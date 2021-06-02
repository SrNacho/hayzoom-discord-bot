from numpy import column_stack, load
import pandas as pd
from datetime import datetime
import xlwt
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


#xls = pd.read_excel("horarios.xls", sheet_name="Hoja 1")
#xls = load_workbook('./horarios.xlsx')
#hoja = xls['Hoja 1']
xls = pd.read_excel("horarios.xlsx", sheet_name="Sheet", )
df = pd.DataFrame(xls.values)
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
ws.title="Hoja 1"
wb.save(filename = 'horarios.xlsx')
#for col in xls["Dia"]:
#     if str(col) == "01/06/2021:01:07":
#        
#         fecha = xls.at[2,"Dia"]
#         fecha_obj = datetime.strptime(fecha,"%d/%m/%Y:%H:%M")
#         xls.at[2,"Dia"] = datetime.strftime(fecha_obj + pd.Timedelta("5 days"),"%d/%m/%Y:%H:%M:%S")
#         print("hola")
# xls = pd.DataFrame(xls)
# print(xls)
# xls.to_excel('horarios.xls', sheet_name='Hoja 1', index=False)

#import schedule, time
#def hola():
#    print("toma")

#schedule.every().day.at("00:56").do(hola)
#while True:
#    schedule.run_pending()
#    time.sleep(1)

